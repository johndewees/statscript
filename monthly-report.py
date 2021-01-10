from openpyxl import load_workbook

filename = 'Digitization Statistics.xlsx'
wb = load_workbook(filename)
ws = wb['Statistics']

month = '2021-01'

cadate = 9
catime = 6
eddate = 12
edtime = 10
mddate = 15
mdtime = 13
qcdate = 18
qctime = 16
uldate = 21
ultime = 19

minimumcol = 1
maximumcol = 23
minimumrow = 2
maximumrow = 10000

cumcatime = 0
cumedtime = 0
cummdtime = 0
cumqctime = 0
cumultime = 0

iterationrow1 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=cadate, max_row=maximumrow, max_col=cadate):
    for cell in row:
        testvar1 = str(ws.cell(row=iterationrow1, column=cadate).value)
        if testvar1 == None:
            continue
        elif testvar1.find(month) != -1:
            catimeval = int(ws.cell(row=iterationrow1, column=catime).value)
            cumcatime = cumcatime + catimeval
    iterationrow1 = iterationrow1 + 1
iterationrow2 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=eddate, max_row=maximumrow, max_col=eddate):
    for cell in row:
        testvar2 = str(ws.cell(row=iterationrow2, column=eddate).value)
        if testvar2 == None:
            continue
        elif testvar2.find(month) != -1:
            edtimeval = int(ws.cell(row=iterationrow2, column=edtime).value)
            cumedtime = cumedtime + edtimeval
    iterationrow2 = iterationrow2 + 1
iterationrow3 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=mddate, max_row=maximumrow, max_col=mddate):
    for cell in row:
        testvar3 = str(ws.cell(row=iterationrow3, column=mddate).value)
        if testvar3 == None:
            continue
        elif testvar3.find(month) != -1:
            mdtimeval = int(ws.cell(row=iterationrow3, column=mdtime).value)
            cummdtime = cummdtime + mdtimeval
    iterationrow3 = iterationrow3 + 1
iterationrow4 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=qcdate, max_row=maximumrow, max_col=qcdate):
    for cell in row:
        testvar4 = str(ws.cell(row=iterationrow4, column=qcdate).value)
        if testvar4 == None:
            continue
        elif testvar4.find(month) != -1:
            qctimeval = int(ws.cell(row=iterationrow4, column=qctime).value)
            cumqctime = cumqctime + qctimeval
    iterationrow4 = iterationrow4 + 1
iterationrow5 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=uldate, max_row=maximumrow, max_col=uldate):
    for cell in row:
        testvar5 = str(ws.cell(row=iterationrow5, column=uldate).value)
        if testvar5 == None:
            continue
        elif testvar5.find(month) != -1:
            ultimeval = int(ws.cell(row=iterationrow5, column=ultime).value)
            cumultime = cumultime + ultimeval
    iterationrow5 = iterationrow5 + 1

print('----------------------------------------------------------')
print('ALL CUMULATIVE CAPTURE TIMES in',month)
print('----------------------------------------------------------')
print('CUMULATIVE CAPTURE TIME IN',month)
print(cumcatime)
print('CUMULATIVE EDIT TIME IN',month)
print(cumedtime)
print('CUMULATIVE METADATA TIME IN',month)
print(cummdtime)
print('CUMULATIVE QUALITY CONROL TIME IN',month)
print(cumqctime)
print('CUMULATIVE UPLOAD TIME IN',month)
print(cumultime)

numobj = 4
numima = 5
capdev = 7
libowned = 3
planetscanitemlib = 0
planetscanimagelib = 0
bookscanitemlib = 0
bookscanimagelib = 0
flatbedscanitemlib = 0
flatbedscanimagelib = 0
planetscanitemnon = 0
planetscanimagenon = 0
bookscanitemnon = 0
bookscanimagenon = 0
flatbedscanitemnon = 0
flatbedscanimagenon = 0
itemsscanned = 0
imagesscanned = 0

iterationrow7 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=cadate, max_row=maximumrow, max_col=cadate):
    for cell in row:
        testvar7 = str(ws.cell(row=iterationrow7, column=cadate).value)
        testvar8 = str(ws.cell(row=iterationrow7, column=capdev).value)
        testvar9 = str(ws.cell(row=iterationrow7, column=libowned).value)
        if testvar7 == None:
            continue
        elif testvar7.find(month) != -1:
            if testvar9.find('Yes') != -1:
                if testvar8.startswith('Planetary'):
                    planetitem1 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    planetscanitemlib = planetscanitemlib + planetitem1
                    planetimage1 = int(ws.cell(row=iterationrow7, column=numima).value)
                    planetscanimagelib = planetscanimagelib + planetimage1
                elif testvar8.startswith('Book'):
                    bookitem1 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    bookscanitemlib = bookscanitemlib + bookitem1
                    bookimage1 = int(ws.cell(row=iterationrow7, column=numima).value)
                    bookscanimagelib = bookscanimagelib + bookimage1
                elif testvar8.startswith('Flatbed'):
                    flatbeditem1 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    flatbedscanitemlib = flatbedscanitemlib + flatbeditem1
                    flatbedimage1 = int(ws.cell(row=iterationrow7, column=numima).value)
                    flatbedscanimagelib = flatbedscanimagelib + flatbedimage1
            elif testvar9.find('No') != -1:
                if testvar8.startswith('Planetary'):
                    planetitem2 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    planetscanitemnon = planetscanitemnon + planetitem2
                    planetimage2 = int(ws.cell(row=iterationrow7, column=numima).value)
                    planetscanimagenon = planetscanimagenon + planetimage2
                elif testvar8.startswith('Book'):
                    bookitem2 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    bookscanitemnon = bookscanitemnon + bookitem2
                    bookimage2 = int(ws.cell(row=iterationrow7, column=numima).value)
                    bookscanimagenon = bookscanimagenon + bookimage2
                elif testvar8.startswith('Flatbed'):
                    flatbeditem2 = int(ws.cell(row=iterationrow7, column=numobj).value)
                    flatbedscanitemnon = flatbedscanitemnon + flatbeditem2
                    flatbedimage2 = int(ws.cell(row=iterationrow7, column=numima).value)
                    flatbedscanimagenon = flatbedscanimagenon + flatbedimage2
    iterationrow7 = iterationrow7 + 1

print('----------------------------------------------------------')
print('ALL LIBRARY OWNED ITEMS SCANNED IN',month)
print('----------------------------------------------------------')
print('BOOK SCANNER ITEMS IN',month)
print(bookscanitemlib)
print('BOOK SCANNER IMAGES IN',month)
print(bookscanimagelib)
print('PLANETARY ITEMS IN',month)
print(planetscanitemlib)
print('PLANETARY IMAGES IN',month)
print(planetscanimagelib)
print('FLATBED ITEMS IN',month)
print(flatbedscanitemlib)
print('FLATBED IMAGES IN',month)
print(flatbedscanimagelib)
print('----------------------------------------------------------')
print('ALL NON-LIBRARY OWNED ITEMS SCANNED IN',month)
print('----------------------------------------------------------')
print('BOOK SCANNER ITEMS IN',month)
print(bookscanitemnon)
print('BOOK SCANNER IMAGES IN',month)
print(bookscanimagenon)
print('PLANETARY ITEMS IN',month)
print(planetscanitemnon)
print('PLANETARY IMAGES IN',month)
print(planetscanimagenon)
print('FLATBED ITEMS IN',month)
print(flatbedscanitemnon)
print('FLATBED IMAGES IN',month)
print(flatbedscanimagenon)

itemsscanned = bookscanitemlib + planetscanitemlib + flatbedscanitemlib + bookscanitemnon + planetscanitemnon + flatbedscanitemnon
imagesscanned = bookscanimagelib + planetscanimagelib + flatbedscanimagelib + bookscanimagenon + planetscanimagenon + flatbedscanimagenon

print('----------------------------------------------------------')
print('GRAND TOTALS FOR SCANNING IN',month)
print('----------------------------------------------------------')
print('TOTAL ITEMS SCANNED IN',month)
print(itemsscanned)
print('TOTAL IMAGES GENERATED IN',month)
print(imagesscanned)

numobj = 4
numima = 5
itemsuploaded = 0
imagesuploaded = 0

iterationrow6 = 2
for row in ws.iter_rows(min_row=minimumrow, min_col=uldate, max_row=maximumrow, max_col=uldate):
    for cell in row:
        testvar6 = str(ws.cell(row=iterationrow6, column=uldate).value)
        if testvar6 == None:
            continue
        elif testvar6.find(month) != -1:
            itemval = int(ws.cell(row=iterationrow6, column=numobj).value)
            itemsuploaded = itemsuploaded + itemval
            imageval = int(ws.cell(row=iterationrow6, column=numima).value)
            imagesuploaded = imagesuploaded + imageval
    iterationrow6 = iterationrow6 + 1

print('----------------------------------------------------------')
print('ALL ITEMS/IMAGES UPLOAD IN',month)
print('----------------------------------------------------------')
print('ITEMS UPLOADED IN',month)
print(itemsuploaded)
print('IMAGES UPLOADED IN',month)
print(imagesuploaded)
print('----------------------------------------------------------')